"""Reporting API tests.

The claim under test is not "a report returns rows" but "a report agrees with
the module it reports on". Each report is therefore compared against the read
endpoint it is built from, and the spreadsheet is compared against the JSON --
a report that quietly diverges from its source is the failure that matters.
"""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import text

from app.core.config import settings
from app.main import app

V1 = settings.API_V1_PREFIX

REPORT_KEYS = [
    "occupancy", "employee", "room-status", "booking", "ticket",
    "housekeeping", "sanitization", "alert", "energy",
]

#: Reports whose date range maps onto a real column. The other three are
#: point-in-time positions and must ignore a range rather than fake one.
DATED = ["booking", "ticket", "housekeeping", "sanitization", "alert", "energy"]
UNDATED = ["occupancy", "employee", "room-status"]

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def client() -> TestClient:
    with TestClient(app) as c:
        yield c


def _sheet(payload: bytes):
    return load_workbook(io.BytesIO(payload)).active


# ---------------------------------------------------------------------------
# Authentication and RBAC -- the existing mechanism, not a new one
# ---------------------------------------------------------------------------


def test_reports_require_a_token(client):
    assert client.get(f"{V1}/reports").status_code == 401


@pytest.mark.parametrize("key", REPORT_KEYS)
def test_every_report_requires_a_token(client, key):
    assert client.get(f"{V1}/reports/{key}").status_code == 401
    assert client.get(f"{V1}/reports/{key}/export.xlsx").status_code == 401


def test_reports_are_gated_on_the_reports_module(client, admin_headers, db):
    """The grant checked must be `reports`, from the seeded registry."""
    granted = db.execute(
        text(
            "SELECT count(*) FROM role_module WHERE module_name = 'reports'"
        )
    ).scalar_one()
    assert granted == 1, "the reports module must exist in role_module"
    assert client.get(f"{V1}/reports", headers=admin_headers).status_code == 200


def test_read_only_role_can_still_read_reports(client, manager_headers):
    """Duty Manager holds reports read-only; reading must not need write."""
    assert client.get(f"{V1}/reports", headers=manager_headers).status_code == 200
    assert client.get(
        f"{V1}/reports/occupancy", headers=manager_headers
    ).status_code == 200


def test_unknown_report_is_404(client, admin_headers):
    r = client.get(f"{V1}/reports/not-a-report", headers=admin_headers)
    assert r.status_code == 404
    assert r.json()["error"]["code"] == "not_found"


# ---------------------------------------------------------------------------
# Definitions -- the screen renders from these, so they must be complete
# ---------------------------------------------------------------------------


def test_all_nine_reports_are_declared(client, admin_headers):
    body = client.get(f"{V1}/reports", headers=admin_headers).json()
    assert [d["key"] for d in body] == REPORT_KEYS


@pytest.mark.parametrize("key", REPORT_KEYS)
def test_each_report_declares_columns_and_a_source(client, admin_headers, key):
    body = client.get(f"{V1}/reports/{key}", headers=admin_headers).json()
    assert body["columns"], f"{key} declares no columns"
    assert body["source"], f"{key} does not name its tables"
    # Column keys must be unique or the table would render one twice.
    keys = [c["key"] for c in body["columns"]]
    assert len(keys) == len(set(keys))


@pytest.mark.parametrize("key", REPORT_KEYS)
def test_rows_carry_exactly_the_declared_columns(client, admin_headers, key):
    """A report must not leak a field it does not declare."""
    body = client.get(
        f"{V1}/reports/{key}", params={"page_size": 5}, headers=admin_headers
    ).json()
    declared = {c["key"] for c in body["columns"]}
    for row in body["items"]:
        assert set(row) == declared, f"{key} row shape drifted from its columns"


# ---------------------------------------------------------------------------
# Each report agrees with the module it reports on
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "key,source_path,source_params",
    [
        ("occupancy", "/occupancy", {}),
        ("room-status", "/occupancy", {}),
        ("employee", "/users", {"is_staff": 1}),
        ("booking", "/stays", {}),
        ("ticket", "/service-requests", {}),
        ("alert", "/alerts", {}),
        ("energy", "/energy-stats", {}),
        ("sanitization", "/maintenance-requests", {"request_type": "disinfection"}),
    ],
)
def test_report_total_matches_its_source_endpoint(
    client, admin_headers, key, source_path, source_params
):
    """The number in the report is the number the module already reports."""
    report = client.get(
        f"{V1}/reports/{key}", params={"page_size": 1}, headers=admin_headers
    ).json()
    source = client.get(
        f"{V1}{source_path}",
        params={"page_size": 1, **source_params},
        headers=admin_headers,
    ).json()
    assert report["total"] == source["total"], (
        f"{key} reports {report['total']} but {source_path} has {source['total']}"
    )


def test_housekeeping_covers_scheduled_and_planned_only(client, admin_headers):
    """Housekeeping is the two non-disinfection tabs; sanitization is the third."""
    rows = client.get(
        f"{V1}/reports/housekeeping", params={"page_size": 100}, headers=admin_headers
    ).json()["items"]
    kinds = {row["maintenance_request_type"] for row in rows}
    assert kinds <= {"scheduled", "planned"}, kinds
    assert "disinfection" not in kinds


def test_employee_report_excludes_guests(client, admin_headers, db):
    """An employee report is staff; a guest is not an employee."""
    total = client.get(
        f"{V1}/reports/employee", params={"page_size": 1}, headers=admin_headers
    ).json()["total"]
    staff = db.execute(
        text("SELECT count(*) FROM app_user WHERE is_staff = 1")
    ).scalar_one()
    assert total == staff


# ---------------------------------------------------------------------------
# Filters -- a filter has to change the result, not just the URL
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("key", DATED)
def test_an_impossible_date_window_returns_nothing(client, admin_headers, key):
    body = client.get(
        f"{V1}/reports/{key}",
        params={"date_from": "1900-01-01", "date_to": "1900-01-02", "page_size": 1},
        headers=admin_headers,
    ).json()
    assert body["total"] == 0, f"{key} ignored its date range"


@pytest.mark.parametrize("key", DATED)
def test_a_date_window_never_widens_the_result(client, admin_headers, key):
    unfiltered = client.get(
        f"{V1}/reports/{key}", params={"page_size": 1}, headers=admin_headers
    ).json()["total"]
    windowed = client.get(
        f"{V1}/reports/{key}",
        params={"date_from": "2026-08-01", "date_to": "2026-08-31", "page_size": 1},
        headers=admin_headers,
    ).json()["total"]
    assert windowed <= unfiltered


@pytest.mark.parametrize("key", UNDATED)
def test_undated_reports_ignore_a_range_and_do_not_claim_it(client, admin_headers, key):
    """These three have no date column. Accepting the parameter silently is
    fine; reporting it as applied would be a lie in the export header."""
    plain = client.get(
        f"{V1}/reports/{key}", params={"page_size": 1}, headers=admin_headers
    ).json()
    ranged = client.get(
        f"{V1}/reports/{key}",
        params={"date_from": "1900-01-01", "date_to": "1900-01-02", "page_size": 1},
        headers=admin_headers,
    ).json()
    assert ranged["total"] == plain["total"]
    assert "date_from" not in ranged["filters_applied"]
    assert "date_to" not in ranged["filters_applied"]


def test_a_filter_the_report_does_not_declare_is_ignored(client, admin_headers):
    """`assigned_to` is a ticket filter, not an occupancy one."""
    plain = client.get(
        f"{V1}/reports/occupancy", params={"page_size": 1}, headers=admin_headers
    ).json()["total"]
    stray = client.get(
        f"{V1}/reports/occupancy",
        params={"page_size": 1, "assigned_to": "00000000-0000-0000-0000-000000000000"},
        headers=admin_headers,
    ).json()["total"]
    assert stray == plain


def test_a_malformed_filter_value_is_rejected(client, admin_headers):
    """Ticket `status` is a service_status id; a word is a 422, not a no-op."""
    r = client.get(
        f"{V1}/reports/ticket", params={"status": "not-an-int"}, headers=admin_headers
    )
    assert r.status_code == 422


def test_status_is_an_int_for_tickets_but_a_label_for_bookings(client, admin_headers):
    """The same parameter name is a different type per report, which is why the
    coercion is driven by each report's own filter declaration."""
    assert client.get(
        f"{V1}/reports/ticket", params={"status": 2, "page_size": 1},
        headers=admin_headers,
    ).status_code == 200
    assert client.get(
        f"{V1}/reports/booking", params={"status": "active", "page_size": 1},
        headers=admin_headers,
    ).status_code == 200


def test_page_size_over_the_cap_is_422(client, admin_headers):
    r = client.get(
        f"{V1}/reports/occupancy", params={"page_size": 1000}, headers=admin_headers
    )
    assert r.status_code == 422


# ---------------------------------------------------------------------------
# Excel export -- the file must be the report, not a second opinion
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("key", REPORT_KEYS)
def test_export_returns_a_real_xlsx_attachment(client, admin_headers, key):
    r = client.get(f"{V1}/reports/{key}/export.xlsx", headers=admin_headers)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(XLSX)
    assert "attachment" in r.headers["content-disposition"]
    assert key.replace("_", "-") in r.headers["content-disposition"]
    # A workbook that will not open is not an export.
    assert _sheet(r.content).max_row >= 6


@pytest.mark.parametrize("key", REPORT_KEYS)
def test_export_headers_are_the_declared_column_headers(client, admin_headers, key):
    body = client.get(f"{V1}/reports/{key}", headers=admin_headers).json()
    sheet = _sheet(client.get(f"{V1}/reports/{key}/export.xlsx", headers=admin_headers).content)
    # Row 6 is the header band; rows 1-4 are the title/meta block.
    written = [c.value for c in sheet[6]][: len(body["columns"])]
    assert written == [c["header"] for c in body["columns"]]


@pytest.mark.parametrize("key", REPORT_KEYS)
def test_export_row_count_matches_the_report_total(client, admin_headers, key):
    """The spreadsheet is not paginated -- it carries every matching row."""
    total = client.get(
        f"{V1}/reports/{key}", params={"page_size": 1}, headers=admin_headers
    ).json()["total"]
    sheet = _sheet(
        client.get(f"{V1}/reports/{key}/export.xlsx", headers=admin_headers).content
    )
    # 5 rows of chrome + 1 header row before the body starts.
    assert sheet.max_row - 6 == total, f"{key} exported {sheet.max_row - 6} of {total}"


def test_export_respects_the_same_filters_as_the_report(client, admin_headers):
    params = {"date_from": "2026-08-17", "date_to": "2026-08-17"}
    total = client.get(
        f"{V1}/reports/alert", params={**params, "page_size": 1}, headers=admin_headers
    ).json()["total"]
    sheet = _sheet(
        client.get(f"{V1}/reports/alert/export.xlsx", params=params, headers=admin_headers).content
    )
    assert sheet.max_row - 6 == total


def test_export_names_its_source_and_filters_in_the_header(client, admin_headers):
    sheet = _sheet(
        client.get(
            f"{V1}/reports/booking/export.xlsx",
            params={"date_from": "2026-08-01", "date_to": "2026-08-31"},
            headers=admin_headers,
        ).content
    )
    assert sheet.cell(row=1, column=1).value == "Booking Report"
    assert "stay" in (sheet.cell(row=2, column=1).value or "")
    filters = sheet.cell(row=3, column=1).value or ""
    assert "2026-08-01" in filters and "2026-08-31" in filters


def test_an_empty_report_still_exports_a_valid_workbook(client, admin_headers):
    """A download that fails on no data is worse than an empty sheet."""
    r = client.get(
        f"{V1}/reports/booking/export.xlsx",
        params={"date_from": "1900-01-01", "date_to": "1900-01-02"},
        headers=admin_headers,
    )
    assert r.status_code == 200
    sheet = _sheet(r.content)
    assert "No data" in (sheet.cell(row=7, column=1).value or "")


# ---------------------------------------------------------------------------
# Reports reflect the live database, not a snapshot
# ---------------------------------------------------------------------------


def test_reports_read_through_to_postgres(client, admin_headers, db):
    """Compared against SQL, so a cached or stale projection would fail."""
    report = client.get(
        f"{V1}/reports/energy", params={"page_size": 1}, headers=admin_headers
    ).json()["total"]
    assert report == db.execute(text("SELECT count(*) FROM energy_stat")).scalar_one()


def test_soft_deleted_maintenance_stays_out_of_the_reports(client, admin_headers, db):
    """The reports inherit the soft-delete rule from the read service rather
    than reimplementing it, so a retired row must not appear."""
    live = db.execute(
        text(
            "SELECT count(*) FROM maintenance_request "
            "WHERE status = 1 AND maintenance_request_type = 'disinfection'"
        )
    ).scalar_one()
    total = client.get(
        f"{V1}/reports/sanitization", params={"page_size": 1}, headers=admin_headers
    ).json()["total"]
    assert total == live
