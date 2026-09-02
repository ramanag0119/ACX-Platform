"""Spreadsheet rendering for the nine reports.

Writes the SAME rows and columns `app.services.reports` serves to the screen,
so a download can never disagree with the table it came from. There is no
report-specific code here: the layout is driven entirely by the `ReportDef`.

PDF export is not implemented yet -- the layout is being taken from a reference
document, and guessing it would mean redoing all nine reports.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.reports import ReportDef

#: Sheet chrome. Kept in one place so every report looks like the same family.
_TITLE_FONT = Font(bold=True, size=14)
_META_FONT = Font(size=9, italic=True, color="555555")
_HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="3EB1C8")
_THIN = Side(style="thin", color="D9D9D9")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_DATE_FMT = "dd-mm-yyyy"
_DATETIME_FMT = "dd-mm-yyyy hh:mm"


def _sheet_title(title: str) -> str:
    """Excel forbids []:*?/\\ in sheet names and caps them at 31 characters."""
    cleaned = "".join("-" if ch in "[]:*?/\\" else ch for ch in title)
    return cleaned[:31] or "Report"


def _cell_value(value: Any, kind: str) -> Any:
    """Native types where Excel benefits, text otherwise.

    Dates and numbers go in unconverted so the recipient can sort and total
    them; booleans become Yes/No because the source columns are 0/1 ints and a
    raw 0 reads as a quantity.
    """
    if value is None:
        return None
    if kind == "boolean":
        if isinstance(value, bool):
            return "Yes" if value else "No"
        if isinstance(value, (int, float)):
            return "Yes" if value else "No"
        return str(value)
    if kind in ("date", "datetime"):
        if isinstance(value, datetime):
            # openpyxl cannot write a tz-aware datetime.
            return value.replace(tzinfo=None)
        if isinstance(value, date):
            return value
        return str(value)
    if kind == "number":
        return value if isinstance(value, (int, float)) else str(value)
    return value if isinstance(value, str) else str(value)


def _describe_filters(definition: ReportDef, filters: dict) -> str:
    """"From Date: 2026-08-01 | Status: Assigned" for the sheet header."""
    labels = {f.name: f.label for f in definition.filters}
    parts = [
        f"{labels.get(name, name)}: {value}"
        for name, value in filters.items()
        if value is not None and value != ""
    ]
    return " | ".join(parts) if parts else "None"


def to_xlsx(
    definition: ReportDef,
    rows: Sequence[dict],
    *,
    filters: dict,
    generated_by: str | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    """Render one report to an .xlsx byte string."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(definition.title)
    columns = list(definition.columns)
    span = max(len(columns), 1)

    # --- header block: what this is, and what produced it -------------------
    sheet.cell(row=1, column=1, value=definition.title).font = _TITLE_FONT
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)

    stamp = (generated_at or datetime.now()).strftime("%d-%m-%Y %H:%M")
    meta = f"Source: {definition.source}    Generated: {stamp}"
    if generated_by:
        meta += f"    By: {generated_by}"
    sheet.cell(row=2, column=1, value=meta).font = _META_FONT
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)

    sheet.cell(
        row=3, column=1, value=f"Filters: {_describe_filters(definition, filters)}"
    ).font = _META_FONT
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=span)

    sheet.cell(row=4, column=1, value=f"Rows: {len(rows)}").font = _META_FONT

    # --- column headers ----------------------------------------------------
    header_row = 6
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column.header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _CELL_BORDER
        sheet.column_dimensions[get_column_letter(index)].width = column.width
    sheet.row_dimensions[header_row].height = 22

    # --- body --------------------------------------------------------------
    for offset, row in enumerate(rows, start=1):
        for index, column in enumerate(columns, start=1):
            cell = sheet.cell(
                row=header_row + offset,
                column=index,
                value=_cell_value(row.get(column.key), column.kind),
            )
            cell.border = _CELL_BORDER
            if column.kind == "date":
                cell.number_format = _DATE_FMT
            elif column.kind == "datetime":
                cell.number_format = _DATETIME_FMT
            if column.kind == "number":
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    if not rows:
        note = sheet.cell(
            row=header_row + 1, column=1, value="No data matched the selected filters."
        )
        note.font = _META_FONT
        sheet.merge_cells(
            start_row=header_row + 1, start_column=1,
            end_row=header_row + 1, end_column=span,
        )

    # Freeze the header so a long report stays readable while scrolling.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(span)}{header_row + max(len(rows), 1)}"
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def filename_for(definition: ReportDef, extension: str) -> str:
    """`occupancy-report-20260825.xlsx` -- stable, sortable, no spaces."""
    slug = definition.key.replace("_", "-")
    return f"{slug}-report-{datetime.now():%Y%m%d}.{extension}"
